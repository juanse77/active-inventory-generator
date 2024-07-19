import sys
import xml.etree.ElementTree as ET
import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def generate_excel(df, excel_file_name):

    with ExcelWriter(excel_file_name, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Actives Inventory', index=False)

    wb = load_workbook(excel_file_name)
    ws = wb['Actives Inventory']

    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    cell_font = Font(name='Arial', size=12)
    cell_alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical='center', horizontal='center')
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for i, cell in enumerate(row):
            if i == 0:
                cell.alignment = Alignment(vertical='center', horizontal='center')
            else:
                cell.alignment = cell_alignment

            cell.fill = PatternFill(start_color='eeeeee', end_color='eeeeee', fill_type='solid')
            cell.font = cell_font
            cell.border = thin_border

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and '\n' in cell.value:
                cell.alignment = cell_alignment
                max_line_count = cell.value.count('\n') + 1
                ws.row_dimensions[cell.row].height = max(15 * max_line_count, ws.row_dimensions[cell.row].height or 15)

    widths = [25, 50, 120]

    for i, col in enumerate(ws.columns):
        column = col[0].column_letter
        ws.column_dimensions[column].width = widths[i]
    
    wb.save(excel_file_name)

def replace_substring(test_str, s1, s2):
    result = ""
    i = 0
    while i < len(test_str):
        if test_str[i:i+len(s1)] == s1:
            result += s2
            i += len(s1)
        else:
            result += test_str[i]
            i += 1
    return result

def get_info_vuln(node, port="unkown", service="unknown"):
    vuln_cve = []
    vulnerabilities = []

    elem = node.find("table")

    if elem is not None:

        cves = elem.find(".//table[@key='ids']") if elem is not None else None

        if cves is not None:
            for cve in cves.findall('elem'):
                if cve.text.split(':')[0] == "CVE":
                    vuln_cve.append(cve.text.split(':')[-1])
                else:
                    vuln_cve.append(cve.text)

        info = {
            "title": elem.find(".//elem[@key='title']").text,
            "state": elem.find(".//elem[@key='state']").text,
            "description": elem.find(".//table[@key='description']/elem").text,
            "disclosure_date": elem.find(".//elem[@key='disclosure']").text if elem.find(".//elem[@key='disclosure']") is not None else 'unknown',
            "references": [e.text for e in elem.findall(".//table[@key='refs']/elem")]
        }

        vulnerabilities.append({
            'port': port,
            'service': service,
            'name': info.get('title'),
            'description': info.get('description'),
            'disclosure_date': info.get('disclosure_date'),
            'references': info.get('references'),
            'cve': vuln_cve
        })

    return vulnerabilities

def extract_host_info(host):
    ip = host.find('address').attrib['addr']
    ports = []
    vulnerabilities = []

    for port in host.find('ports').findall('port'):
        port_id = port.attrib['portid']
        service_node = port.find('service')
        service = service_node.attrib.get('name', 'unknown') if service_node is not None else 'unknown'
        ports.append((port_id, service))

        for script in port.findall('script'):
            if 'vulnerable' in script.attrib['output'].lower():
                vulnerabilities.extend(get_info_vuln(script, port_id, service))
                
    hostscript = host.find('hostscript')
    if hostscript is not None:
        for script in hostscript.findall('script'):
            if 'vulnerable' in script.attrib['output'].lower():
                vulnerabilities.extend(get_info_vuln(script))

    return ip, ports, vulnerabilities

def main(xml_file, xlsx_file_name):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    hosts_info = []
    for host in root.findall('host'):
        ip, ports, vulnerabilities = extract_host_info(host)
        hosts_info.append({
            'ip': ip,
            'ports': ports,
            'vulnerabilities': vulnerabilities
        })

    for host_info in hosts_info:
        print(f"IP Direction: {host_info['ip']}")
        print("Ports and Services:")
        for port, service in host_info['ports']:
            print(f"  - Port: {port}, Service: {service}")
        print("Vulnerabilities:")
        for vuln in host_info['vulnerabilities']:
            description = replace_substring(vuln['description'], "\n", "\n\t")
            formated_references = '\n\t'.join(vuln['references'])
            print(f"  - Port: {vuln['port']}, Service: {vuln['service']}")
            print(f"    Name: {vuln['name']}")
            print(f"    CVEs: {', '.join(vuln['cve'])}")
            print(f"    Disclosure date: {vuln['disclosure_date']}")
            print(f"    References: \n\t{formated_references}")
            print(f"    Description:\n\t{description}")
            
        print("\n")

    ips = []
    ports_services = []
    vulns = []
    for host_info in hosts_info:
        ports_servs = []
        vuls = []

        ips.append(host_info['ip'])
        for port, service in host_info['ports']:
            ports_servs.append(f"  - Port: {port}, Service: {service}")    
    
        ports_services.append("\n".join(ports_servs))

        for vuln in host_info['vulnerabilities']:
            description = replace_substring(vuln['description'], "\n", "\n\t")
            formated_references = '\n\t'.join(vuln['references'])
            v = f"  - Port: {vuln['port']}, Service: {vuln['service']}\n"
            v += f"    Name: {vuln['name']}\n"
            v += f"    CVEs: {', '.join(vuln['cve'])}\n"
            #v += f"    Disclosure Date: {vuln['disclosure_date']}\n"
            v += f"    References: \n\t{formated_references}\n"
            #v += f"    Description:\n\t{description}\n"

            vuls.append(v)

        vulns.append("\n".join(vuls))

    datos = {
        "IP Direction": ips,
        "Ports and Services": ports_services,
        "Vulnerabilities": vulns
    }

    df = pd.DataFrame(datos)
    
    if not xlsx_file_name.endswith('.xlsx'):
        xlsx_file_name = f"{xlsx_file_name}.xlsx"

    generate_excel(df, xlsx_file_name)

    print("An excel file has been generated with the results\n")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Use: python active-inventory-generator.py <xml-file> <xlsx-file-name>")
        sys.exit(1)

    xml_file = sys.argv[1]
    xlsx_file_name = sys.argv[2]
    main(xml_file, xlsx_file_name)
